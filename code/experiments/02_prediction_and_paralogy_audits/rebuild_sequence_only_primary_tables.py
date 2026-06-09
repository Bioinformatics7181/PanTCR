#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rebuild sequence-only primary TRBV benchmark tables.

This script updates the overall primary metric tables so that the main
sequence-recovery counts ignore upstream gene-label assignment. Per-gene audit
tables remain gene-aware by design.
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


PACKAGE_ROOT = Path(__file__).resolve().parents[3]
EXPERIMENT_DIR = PACKAGE_ROOT / "experiments" / "02_prediction_and_paralogy_audits"
DEFAULT_SOURCE_ROOT = EXPERIMENT_DIR / "generated" / "full_provenance_inputs"
DEFAULT_OUT_DIR = EXPERIMENT_DIR / "generated" / "sequence_only_primary_metrics"
DEFAULT_CLASSIFIED_TRUTH = (
    EXPERIMENT_DIR
    / "generated"
    / "observed_region_recovery"
    / "s5_s6_observed_region_three_class_per_truth.csv"
)


METHOD_MAP = {
    "Bayes": "PanTCR",
    "BayesNoPrior": "PanTCR-NP",
    "FindAlleles": "FindAlleles",
    "MiXCR": "MiXCR-default",
    "PanTCR.semi": "PanTCR",
}

SCENARIO_LABELS = {
    "expr_ScenarioA": "Scenario A",
    "expr_ScenarioB": "Scenario B",
    "expr_ScenarioC": "Scenario C",
    "expr_FullLength": "High-quality full-length",
}

SCENARIO_DIRS = {
    "expr_ScenarioA": "scenario_a",
    "expr_ScenarioB": "scenario_b",
    "expr_ScenarioC": "scenario_c",
    "expr_FullLength": "full_length",
}

PREDICTION_CLASS_COLUMNS = [
    "Exact Seq & \nTrue Gene",
    "Exact Seq & \nWrong Gene",
    "Wrong Seq & \nTruth Gene",
    "Unsupported Prediction",
]


def clean_method(value: object) -> str:
    return METHOD_MAP.get(str(value), str(value))


def normalize_sequence(series: pd.Series) -> pd.Series:
    return series.astype(str).str.upper().str.strip()


def load_truth_prediction_sources(source_root: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    truth_parts: list[pd.DataFrame] = []
    pred_parts: list[pd.DataFrame] = []

    count_root = source_root / "01_count_matrix_and_coverage_strata"
    for expr, subdir in SCENARIO_DIRS.items():
        truth = pd.read_csv(count_root / subdir / "per_truth_call_status.csv", low_memory=False)
        pred = pd.read_csv(count_root / subdir / "per_prediction_status.csv", low_memory=False)
        for df in (truth, pred):
            df["Benchmark"] = "In silico simulation"
            df["Dataset"] = SCENARIO_LABELS[expr]
            df["Method"] = df["method"].map(clean_method)
            df["Sample key"] = df.apply(
                lambda row: f"{row['expr']}|{row['population']}|{row['seed']}|{row['gene_type']}",
                axis=1,
            )
        truth_parts.append(truth[truth["gene_type"].eq("V")].copy())
        pred_parts.append(pred[pred["gene_type"].eq("V")].copy())

    truth = pd.read_csv(
        source_root / "12_semi_simu_evidence_analysis" / "results" / "per_truth_call_status.csv",
        low_memory=False,
    )
    pred = pd.read_csv(
        source_root / "12_semi_simu_evidence_analysis" / "results" / "per_prediction_status.csv",
        low_memory=False,
    )
    for df in (truth, pred):
        df["Benchmark"] = "Semi-synthetic AIRR-seq"
        df["Dataset"] = df["dataset"].astype(str)
        df["Method"] = df["method"].map(clean_method)
        df["Sample key"] = df.apply(
            lambda row: f"{row['expr']}|{row['dataset']}|{row['sample_id']}|{row['gene_type']}",
            axis=1,
        )
    truth_parts.append(truth[truth["gene_type"].eq("V")].copy())
    pred_parts.append(pred[pred["gene_type"].eq("V")].copy())

    truth = pd.read_csv(
        source_root / "13_scbulk_real_evidence_analysis" / "results" / "per_truth_call_status.csv",
        low_memory=False,
    )
    pred = pd.read_csv(
        source_root / "13_scbulk_real_evidence_analysis" / "results" / "per_prediction_status.csv",
        low_memory=False,
    )
    for df in (truth, pred):
        df["Benchmark"] = "Pseudo-bulk RNA-seq"
        df["Dataset"] = df["SC_ID"].astype(str)
        df["Method"] = df["method"].map(clean_method)
        df["Sample key"] = df.apply(
            lambda row: f"{row['SC_ID']}|{row['DatasetID']}|{row['gene_type']}",
            axis=1,
        )
    truth_parts.append(truth[truth["gene_type"].eq("V")].copy())
    pred_parts.append(pred[pred["gene_type"].eq("V")].copy())

    truth_all = pd.concat(truth_parts, ignore_index=True, sort=False)
    pred_all = pd.concat(pred_parts, ignore_index=True, sort=False)
    truth_all["truth_seq"] = normalize_sequence(truth_all["truth_seq"])
    pred_all["pred_seq"] = normalize_sequence(pred_all["pred_seq"])
    truth_all["gene"] = truth_all["gene"].astype(str)
    pred_all["gene"] = pred_all["gene"].astype(str)
    return truth_all, pred_all


def annotate_sequence_only_matches(truth: pd.DataFrame, pred: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Mark one-to-one exact sequence matches while ignoring gene labels.

    Same-gene exact matches are assigned first for interpretability. Remaining
    same-sequence matches are assigned across genes. Extra duplicate predictions
    with an already matched sequence remain unmatched and are counted as extra
    predictions in the primary count matrix.
    """
    truth = truth.copy()
    pred = pred.copy()
    truth["Sequence-only exact matched"] = False
    pred["Sequence-only exact matched"] = False
    pred_groups = {
        key: group
        for key, group in pred.groupby(["Benchmark", "Dataset", "Method", "Sample key"], sort=False)
    }
    for key, truth_group in truth.groupby(["Benchmark", "Dataset", "Method", "Sample key"], sort=False):
        pred_group = pred_groups.get(key)
        if pred_group is None or pred_group.empty:
            continue
        used_truth: set[int] = set()
        used_pred: set[int] = set()
        for seq, pred_indices in pred_group.groupby("pred_seq", sort=False).groups.items():
            truth_indices = truth_group.index[truth_group["truth_seq"].eq(seq)].tolist()
            if not truth_indices:
                continue
            for pred_idx in list(pred_indices):
                if pred_idx in used_pred:
                    continue
                pred_gene = pred.at[pred_idx, "gene"]
                candidates = [
                    truth_idx
                    for truth_idx in truth_indices
                    if truth_idx not in used_truth and truth.at[truth_idx, "gene"] == pred_gene
                ]
                if candidates:
                    truth_idx = candidates[0]
                    used_pred.add(pred_idx)
                    used_truth.add(truth_idx)
                    pred.at[pred_idx, "Sequence-only exact matched"] = True
                    truth.at[truth_idx, "Sequence-only exact matched"] = True
            for pred_idx in list(pred_indices):
                if pred_idx in used_pred:
                    continue
                candidates = [
                    truth_idx
                    for truth_idx in truth_indices
                    if truth_idx not in used_truth and truth.at[truth_idx, "gene"] != pred.at[pred_idx, "gene"]
                ]
                if candidates:
                    truth_idx = candidates[0]
                    used_pred.add(pred_idx)
                    used_truth.add(truth_idx)
                    pred.at[pred_idx, "Sequence-only exact matched"] = True
                    truth.at[truth_idx, "Sequence-only exact matched"] = True
    return truth, pred


def build_sequence_only_count_matrix(truth: pd.DataFrame, pred: pd.DataFrame) -> pd.DataFrame:

    rows: list[dict[str, object]] = []
    for (benchmark, dataset, method), truth_group in truth.groupby(["Benchmark", "Dataset", "Method"], sort=False):
        pred_group = pred[
            pred["Benchmark"].eq(benchmark)
            & pred["Dataset"].eq(dataset)
            & pred["Method"].eq(method)
        ]
        truth_n = len(truth_group)
        pred_n = len(pred_group)
        tp = int(truth_group["Sequence-only exact matched"].sum())
        matched_pred = int(pred_group["Sequence-only exact matched"].sum())
        fp = pred_n - matched_pred
        fn = truth_n - tp
        rows.append(
            {
                "Benchmark": benchmark,
                "Dataset": dataset,
                "Method": method,
                "NO. of Truth Alleles": truth_n,
                "NO. of Predicted Alleles": pred_n,
                "TP": tp,
                "FP": fp,
                "FN": fn,
            }
        )

    out = pd.DataFrame(rows)
    out = append_pseudo_total(out, ["NO. of Truth Alleles", "NO. of Predicted Alleles", "TP", "FP", "FN"])
    return out


def build_prediction_gene_label_audit(pred: pd.DataFrame, truth: pd.DataFrame) -> pd.DataFrame:
    truth_lookup: dict[tuple[str, str, str, str], tuple[set[tuple[str, str]], set[str], set[str]]] = {}
    for key, group in truth.groupby(["Benchmark", "Dataset", "Method", "Sample key"], sort=False):
        truth_lookup[key] = (
            set(zip(group["gene"], group["truth_seq"])),
            set(group["truth_seq"]),
            set(group["gene"]),
        )

    classes: list[str] = []
    for _, row in pred.iterrows():
        key = (row["Benchmark"], row["Dataset"], row["Method"], row["Sample key"])
        gene_seq, seq_set, gene_set = truth_lookup.get(key, (set(), set(), set()))
        pred_key = (str(row["gene"]), str(row["pred_seq"]))
        if pred_key in gene_seq:
            classes.append("Exact Seq & \nTrue Gene")
        elif str(row["pred_seq"]) in seq_set:
            classes.append("Exact Seq & \nWrong Gene")
        elif str(row["gene"]) in gene_set:
            classes.append("Wrong Seq & \nTruth Gene")
        else:
            classes.append("Unsupported Prediction")

    pred = pred.copy()
    pred["Prediction class"] = classes
    matrix = (
        pred.groupby(["Benchmark", "Dataset", "Method", "Prediction class"], dropna=False)
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    for col in PREDICTION_CLASS_COLUMNS:
        if col not in matrix.columns:
            matrix[col] = 0
    matrix["NO. of Predicted Alleles"] = matrix[PREDICTION_CLASS_COLUMNS].sum(axis=1)
    matrix = matrix[
        ["Benchmark", "Dataset", "Method", "NO. of Predicted Alleles", *PREDICTION_CLASS_COLUMNS]
    ]
    matrix = append_pseudo_total(matrix, ["NO. of Predicted Alleles", *PREDICTION_CLASS_COLUMNS])
    return matrix


def build_truth_recovery_table(classified_truth_csv: Path, pred: pd.DataFrame) -> pd.DataFrame:
    truth = pd.read_csv(classified_truth_csv, low_memory=False)
    truth = truth[truth["gene_type"].eq("V")].copy()
    truth["Method"] = truth["Method"].map(clean_method)
    truth["truth_seq"] = normalize_sequence(truth["truth_seq"])

    def sample_key(row: pd.Series) -> str:
        if row["Benchmark"] == "In silico simulation":
            return f"{row['expr']}|{row['population']}|{row['seed']}|{row['gene_type']}"
        if row["Benchmark"] == "Semi-synthetic AIRR-seq":
            return f"{row['expr']}|{row['dataset']}|{row['sample_id']}|{row['gene_type']}"
        return f"{row['SC_ID']}|{row['DatasetID']}|{row['gene_type']}"

    truth["Sample key"] = truth.apply(sample_key, axis=1)
    truth, _ = annotate_sequence_only_matches(truth, pred)
    truth["Sequence-exact recovery"] = truth["Sequence-only exact matched"]
    truth["New recovery class"] = truth.apply(
        lambda row: "Exact Recovery"
        if row["Sequence-exact recovery"]
        else (
            "Observed-region Compatible Partial Recovery"
            if row["Observed-region recovery class"] == "Observed-region Compatible Partial Recovery"
            else "Not Recovered or Incompatible"
        ),
        axis=1,
    )

    rows = []
    keep_methods = {
        "In silico simulation": {"MiXCR-default", "FindAlleles", "PanTCR-NP", "PanTCR"},
        "Semi-synthetic AIRR-seq": {"MiXCR-default", "FindAlleles", "PanTCR"},
        "Pseudo-bulk RNA-seq": {"MiXCR-default", "FindAlleles", "PanTCR"},
    }
    for (benchmark, dataset, method), group in truth.groupby(["Benchmark", "Dataset", "Method"], sort=False):
        if method not in keep_methods.get(benchmark, set()):
            continue
        exact = int((group["New recovery class"] == "Exact Recovery").sum())
        partial = int((group["New recovery class"] == "Observed-region Compatible Partial Recovery").sum())
        not_recovered = int((group["New recovery class"] == "Not Recovered or Incompatible").sum())
        total = exact + partial + not_recovered
        rows.append(
            {
                "Benchmark": benchmark,
                "Dataset": dataset,
                "Method": method,
                "NO. of Truth Alleles": total,
                "Exact Recovery": exact,
                "Observed-region Compatible Partial Recovery": partial,
                "Not Recovered or Incompatible": not_recovered,
                "Exact Rate": exact / total if total else 0.0,
                "Exact or Partial Rate": (exact + partial) / total if total else 0.0,
            }
        )
    out = pd.DataFrame(rows)
    out = append_pseudo_total(
        out,
        [
            "NO. of Truth Alleles",
            "Exact Recovery",
            "Observed-region Compatible Partial Recovery",
            "Not Recovered or Incompatible",
        ],
    )
    out["Exact Rate"] = out["Exact Recovery"] / out["NO. of Truth Alleles"]
    out["Exact or Partial Rate"] = (
        out["Exact Recovery"] + out["Observed-region Compatible Partial Recovery"]
    ) / out["NO. of Truth Alleles"]
    return out


def append_pseudo_total(df: pd.DataFrame, sum_cols: list[str]) -> pd.DataFrame:
    pieces = [df]
    pseudo = df[df["Benchmark"].eq("Pseudo-bulk RNA-seq") & ~df["Dataset"].eq("Total")].copy()
    if not pseudo.empty:
        totals = (
            pseudo.groupby(["Benchmark", "Method"], dropna=False)[sum_cols]
            .sum()
            .reset_index()
        )
        totals.insert(1, "Dataset", "Total")
        for col in df.columns:
            if col not in totals.columns:
                totals[col] = pd.NA
        pieces.append(totals[df.columns])
    return pd.concat(pieces, ignore_index=True, sort=False)


def update_sheet_from_table(
    workbook_path: Path,
    output_path: Path,
    sheet_name: str,
    table: pd.DataFrame,
    columns: list[str],
) -> None:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    lookup = {
        (row["Benchmark"], row["Dataset"], row["Method"]): row
        for _, row in table.iterrows()
    }
    benchmark = dataset = None
    for row_idx in range(3, ws.max_row + 1):
        first = ws.cell(row_idx, 1).value
        second = ws.cell(row_idx, 2).value
        method = ws.cell(row_idx, 3).value
        if first:
            benchmark = first
        if second:
            dataset = second
        if not method or str(method).startswith("Note:"):
            continue
        key = (benchmark, dataset, method)
        if key not in lookup:
            continue
        record = lookup[key]
        for offset, col_name in enumerate(columns, start=4):
            value = record[col_name]
            if col_name.endswith("Rate"):
                value = f"{float(value):.3f}"
            ws.cell(row_idx, offset).value = value
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_outputs(
    workbook: Path | None,
    output_workbook: Path | None,
    out_dir: Path,
    s2: pd.DataFrame,
    s5: pd.DataFrame,
    s7: pd.DataFrame,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    s2.to_csv(out_dir / "supplementary_table_s2_sequence_only.csv", index=False)
    s5.to_csv(out_dir / "supplementary_table_s5_sequence_only.csv", index=False)
    s7.to_csv(out_dir / "supplementary_table_s7_sequence_gene_label_audit.csv", index=False)

    if workbook is None or output_workbook is None:
        return

    update_sheet_from_table(
        workbook,
        output_workbook,
        "Supplementary Table S2",
        s2,
        ["NO. of Truth Alleles", "NO. of Predicted Alleles", "TP", "FP", "FN"],
    )
    update_sheet_from_table(
        output_workbook,
        output_workbook,
        "Supplementary Table S5",
        s5,
        [
            "NO. of Truth Alleles",
            "Exact Recovery",
            "Observed-region Compatible Partial Recovery",
            "Not Recovered or Incompatible",
            "Exact Rate",
            "Exact or Partial Rate",
        ],
    )
    update_sheet_from_table(
        output_workbook,
        output_workbook,
        "Supplementary Table S7",
        s7,
        ["NO. of Predicted Alleles", *PREDICTION_CLASS_COLUMNS],
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-root", default=str(DEFAULT_SOURCE_ROOT))
    parser.add_argument(
        "--classified-truth",
        default=str(DEFAULT_CLASSIFIED_TRUTH),
        help="Observed-region three-class per-truth CSV written by summarize_observed_region_recovery_s5_s6.py --write-per-truth.",
    )
    parser.add_argument(
        "--workbook",
        default="",
        help="Optional Supplementary Tables workbook template to update. If omitted, only CSV outputs are written.",
    )
    parser.add_argument(
        "--output-workbook",
        default="",
        help="Optional output workbook path. Required only when --workbook is supplied.",
    )
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_root = Path(args.source_root)
    if not source_root.exists():
        raise FileNotFoundError(
            f"Full-provenance source root not found: {source_root}. "
            "Provide --source-root or set FULL_PROVENANCE_ROOT in run_rebuild.sh."
        )
    truth, pred = load_truth_prediction_sources(source_root)
    matched_truth, matched_pred = annotate_sequence_only_matches(truth, pred)
    s2 = build_sequence_only_count_matrix(matched_truth, matched_pred)

    classified_truth = Path(args.classified_truth)
    if not classified_truth.exists():
        raise FileNotFoundError(classified_truth)
    s5 = build_truth_recovery_table(classified_truth, pred)
    s7 = build_prediction_gene_label_audit(pred, truth)
    workbook = Path(args.workbook) if args.workbook else None
    output_workbook = Path(args.output_workbook) if args.output_workbook else None
    if workbook is not None and output_workbook is None:
        raise ValueError("--output-workbook is required when --workbook is supplied.")
    if workbook is not None and not workbook.exists():
        raise FileNotFoundError(workbook)
    write_outputs(workbook, output_workbook, Path(args.out_dir), s2, s5, s7)
    if output_workbook is not None:
        print(output_workbook)
    print(Path(args.out_dir))


if __name__ == "__main__":
    main()
