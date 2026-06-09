#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build three-class observed-region recovery versions of Supplementary S5/S6.

This script does not modify the frozen supplementary workbook.
The incomplete-recovery class is based on MiXCR-derived observed regions
extracted by code/pantcr/collect_mutation.py before PanTCR NaiveDiversityIndex inference
filtering.
"""

from __future__ import annotations

import importlib.util
import argparse
import math
import sys
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CODE_EXPERIMENT_DIR = Path(__file__).resolve().parent
PACKAGE_ROOT = CODE_EXPERIMENT_DIR.parents[2]
EXPERIMENT_DIR = PACKAGE_ROOT / "experiments" / CODE_EXPERIMENT_DIR.name
DEFAULT_OUT_DIR = EXPERIMENT_DIR / "generated" / "observed_region_recovery"
MATRIX_SCRIPT = CODE_EXPERIMENT_DIR / "build_gene_level_recovery_matrices.py"

DATASET_LABELS = {
    "expr_ScenarioA": "Scenario A",
    "expr_ScenarioB": "Scenario B",
    "expr_ScenarioC": "Scenario C",
    "expr_FullLength": "High-quality full-length",
}

CLASS_EXACT = "Exact recovery"
CLASS_PARTIAL = "Observed-region Compatible Partial Recovery"
CLASS_NOT = "Not Recovered / Incompatible"


def load_matrix_module():
    spec = importlib.util.spec_from_file_location("gene_level_recovery_matrices", MATRIX_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load script: {MATRIX_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def sample_key(row: pd.Series, cols: Iterable[str]) -> tuple[str, ...]:
    return tuple(str(row.get(col, "")) for col in cols)


def parse_ranges_part(part: object) -> list[tuple[int, int]]:
    if part is None or pd.isna(part):
        return []
    out: list[tuple[int, int]] = []
    for seg in str(part).split(";"):
        seg = seg.strip()
        if not seg:
            continue
        toks = seg.split(",")
        if len(toks) != 2:
            continue
        try:
            a, b = int(toks[0]), int(toks[1])
        except ValueError:
            continue
        if b < a:
            a, b = b, a
        out.append((a, b))
    return out


def observed_positions_by_gene(evidence_csv: object) -> dict[str, set[int]]:
    """Read raw collect_mutation.py ObservedRanges without min_naive filtering."""
    path = Path(str(evidence_csv))
    if not path.exists():
        return {}
    df = pd.read_csv(path, low_memory=False)
    if "Family" not in df.columns or "ObservedRanges" not in df.columns:
        raise ValueError(f"Mutation evidence file missing Family/ObservedRanges: {path}")
    out: dict[str, set[int]] = {}
    for row in df.itertuples(index=False):
        gene = str(getattr(row, "Family", "")).strip()
        obs = getattr(row, "ObservedRanges", "")
        if not gene:
            continue
        positions = out.setdefault(gene, set())
        for piece in str(obs).split("|"):
            for a, b in parse_ranges_part(piece):
                positions.update(range(a, b))
    return out


def build_prediction_index(pred: pd.DataFrame, sample_cols: tuple[str, ...]) -> dict[tuple[str, ...], list[str]]:
    out: dict[tuple[str, ...], list[str]] = {}
    for _, row in pred.iterrows():
        gene = str(row.get("gene", ""))
        seq = str(row.get("pred_seq", "")).strip().upper()
        if not gene or not seq:
            continue
        out.setdefault(sample_key(row, sample_cols) + (gene,), []).append(seq)
    return out


def sequence_matches_positions(pred_seq: str, truth_seq: str, positions: set[int]) -> bool:
    pred_seq = str(pred_seq).strip().upper()
    truth_seq = str(truth_seq).strip().upper()
    if not positions:
        return False
    for pos in positions:
        if pos < 0 or pos >= len(truth_seq):
            continue
        if pos >= len(pred_seq) or pred_seq[pos] != truth_seq[pos]:
            return False
    return True


def classify_truth_rows_for_spec(spec, truth: pd.DataFrame, pred: pd.DataFrame) -> pd.DataFrame:
    pred_index = build_prediction_index(pred, spec.sample_cols)
    observed_cache: dict[str, dict[str, set[int]]] = {}
    classes = []
    observed_position_counts = []

    for _, row in truth.iterrows():
        if str(row.get("status", "")) == "exact_tp":
            classes.append(CLASS_EXACT)
            observed_position_counts.append(math.nan)
            continue

        gene = str(row.get("gene", ""))
        truth_seq = str(row.get("truth_seq", "")).strip().upper()
        evidence_file = str(row.get("evidence_file", ""))
        if evidence_file not in observed_cache:
            observed_cache[evidence_file] = observed_positions_by_gene(evidence_file)
        raw_positions = observed_cache[evidence_file].get(gene, set())
        positions = {p for p in raw_positions if 0 <= p < len(truth_seq)}
        observed_position_counts.append(len(positions))

        same_gene_predictions = pred_index.get(sample_key(row, spec.sample_cols) + (gene,), [])
        if positions and same_gene_predictions and any(
            sequence_matches_positions(pred_seq, truth_seq, positions) for pred_seq in same_gene_predictions
        ):
            classes.append(CLASS_PARTIAL)
        else:
            classes.append(CLASS_NOT)

    out = truth.copy()
    out["Observed-region recovery class"] = classes
    out["Observed positions from collect_mutation.py"] = observed_position_counts
    return out


def build_classified_truth(source_root: str | None = None) -> pd.DataFrame:
    matrix_module = load_matrix_module()
    if source_root:
        matrix_module.configure_source_root(source_root)
    pieces = []
    for spec in matrix_module.build_dataset_specs():
        truth, pred = matrix_module.normalize_tables(spec)
        classified = classify_truth_rows_for_spec(spec, truth, pred)
        pieces.append(classified)
    out = pd.concat(pieces, ignore_index=True, sort=False)
    out["Dataset"] = out["Dataset"].replace(DATASET_LABELS)
    return out


def add_group_rows(df: pd.DataFrame, add_pseudo_total: bool = True) -> pd.DataFrame:
    pieces = [df]
    all_gene = df.copy()
    all_gene["Gene category"] = "All TRBV"
    pieces.append(all_gene)
    if add_pseudo_total:
        pseudo = df[df["Benchmark"].eq("Pseudo-bulk RNA-seq")].copy()
        if not pseudo.empty:
            pseudo["Dataset"] = "Total"
            pieces.append(pseudo)
            pseudo_all_gene = pseudo.copy()
            pseudo_all_gene["Gene category"] = "All TRBV"
            pieces.append(pseudo_all_gene)
    return pd.concat(pieces, ignore_index=True, sort=False)


def build_lookup(classified: pd.DataFrame) -> pd.DataFrame:
    expanded = add_group_rows(classified)
    group_cols = ["Benchmark", "Dataset", "Method", "Gene category"]
    matrix = (
        expanded.groupby(group_cols + ["Observed-region recovery class"], dropna=False)
        .size()
        .reset_index(name="n")
        .pivot_table(index=group_cols, columns="Observed-region recovery class", values="n", fill_value=0, aggfunc="sum")
        .reset_index()
    )
    matrix.columns.name = None
    for col in [CLASS_EXACT, CLASS_PARTIAL, CLASS_NOT]:
        if col not in matrix.columns:
            matrix[col] = 0
        matrix[col] = matrix[col].astype(int)
    matrix["Recovery Class Sum"] = matrix[[CLASS_EXACT, CLASS_PARTIAL, CLASS_NOT]].sum(axis=1)
    return matrix


def make_output_table(lookup: pd.DataFrame, is_gene: bool) -> pd.DataFrame:
    rows = []
    source = lookup.copy()
    if is_gene:
        source = source[source["Gene category"].astype(str).ne("All TRBV")]
    else:
        source = source[source["Gene category"].astype(str).eq("All TRBV")]
    for _, rec in source.iterrows():
        truth_n = int(rec["Recovery Class Sum"])
        exact = int(rec[CLASS_EXACT])
        partial = int(rec[CLASS_PARTIAL])
        not_recovered = int(rec[CLASS_NOT])
        out = {
            "Benchmark": rec["Benchmark"],
            "Dataset": rec["Dataset"],
            "Method": rec["Method"],
        }
        if is_gene:
            out["Gene"] = rec["Gene category"]
        out.update(
            {
                "No. of Truth Alleles": truth_n,
                CLASS_EXACT: exact,
                CLASS_PARTIAL: partial,
                CLASS_NOT: not_recovered,
                "Recovery Class Sum": exact + partial + not_recovered,
                "Exact Rate": f"{exact / truth_n:.3f}" if truth_n else "0.000",
                "Exact or Observed-region Compatible Partial Rate": f"{(exact + partial) / truth_n:.3f}" if truth_n else "0.000",
            }
        )
        rows.append(out)
    return pd.DataFrame(rows)


def write_df(ws, df: pd.DataFrame, title: str) -> None:
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.cell(1, 1).alignment = Alignment(wrap_text=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(2, c)
        cell.value = col
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(df.itertuples(index=False), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for c_idx, col in enumerate(df.columns, 1):
        width = min(max(len(str(col)) + 4, 12), 42)
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.freeze_panes = "A3"


def add_notes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Definitions")
    rows = [
        ["Term", "Definition"],
        [CLASS_EXACT, "The predicted mature TRBV sequence exactly matches the truth allele in the evaluated V-region."],
        [
            CLASS_PARTIAL,
            (
                "The complete sequence is not exactly recovered, but the prediction is concordant with the truth at all "
                "retained observed positions within the evaluated region. Disagreements, if any, are confined to positions "
                "without retained direct evidence."
            ),
        ],
        [
            CLASS_NOT,
            (
                "All remaining truth alleles, including observed-region conflicts, no same-gene compatible prediction, no-call, "
                "or cases where the upstream gene assignment leaves no usable observed region for the truth gene."
            ),
        ],
        [
            "Table note",
            (
                "Exact recovery denotes a complete match between the predicted and truth allele sequence within the evaluated "
                "mature TRBV region. Observed-region compatible partial recovery denotes cases where the complete allele "
                "sequence was not exactly recovered, but all MiXCR-derived sample-observed positions within the evaluated region "
                "were concordant between the prediction and the truth allele; any disagreement was confined to positions "
                "without retained direct evidence. Observed regions were defined from the MiXCR-derived gene assignment and "
                "coverage mask used during evidence extraction. This category indicates compatibility with observed-region "
                "evidence and should not be interpreted as direct recovery of unobserved allele-defining positions. All "
                "remaining cases, including observed-region conflicts, no compatible same-gene prediction, no-call, or absence "
                "of usable observed evidence due to upstream gene assignment, were classified as not recovered / incompatible."
            ),
        ],
        [
            "Observed region implementation",
            (
                "Observed regions are calculated from the union of `ObservedRanges` in the collect_mutation.py mutation-evidence "
                "CSV for the same sample and TRBV gene before applying the PanTCR NaiveDiversityIndex/min_naive inference filter. "
                "They are MiXCR-derived clone/rearrangement-table evidence ranges projected onto gene-specific reference coordinates, "
                "not raw FASTQ pileups or phased truth-allele coverage maps."
            ),
        ],
    ]
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_idx == 1:
                cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 120


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    parser.add_argument(
        "--write-per-truth",
        action="store_true",
        help="Also write the large per-truth classification CSV. This is off by default for the GitHub-light package.",
    )
    parser.add_argument(
        "--source-root",
        default=None,
        help=(
            "Full provenance root containing the normalized per-truth/per-prediction inputs. "
            "If omitted, PANTCR_FULL_PROVENANCE_ROOT is used by build_gene_level_recovery_matrices.py."
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    out_dir = Path(args.out_dir)
    out_workbook = out_dir / "Supplementary Tables S5-S6 observed-region three-class recovery.xlsx"
    classified = build_classified_truth(args.source_root)
    lookup = build_lookup(classified)

    s5 = make_output_table(lookup, is_gene=False)
    s6 = make_output_table(lookup, is_gene=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Supplementary Table S5"
    write_df(
        ws,
        s5,
        "Supplementary Table S5. Observed-region TRBV allele recovery across evaluated datasets.",
    )
    ws2 = wb.create_sheet("Supplementary Table S6")
    write_df(
        ws2,
        s6,
        "Supplementary Table S6. Per-gene observed-region TRBV allele recovery across evaluated datasets.",
    )
    add_notes_sheet(wb)
    out_workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_workbook)

    if args.write_per_truth:
        matrix_module = load_matrix_module()
        if args.source_root:
            matrix_module.configure_source_root(args.source_root)
        matrix_module.sanitize_source_columns(classified).to_csv(
            out_dir / "s5_s6_observed_region_three_class_per_truth.csv",
            index=False,
        )
    print(out_workbook)


if __name__ == "__main__":
    main()
